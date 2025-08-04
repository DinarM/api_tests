import random
import uuid
from datetime import datetime, timedelta
from http import HTTPStatus
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from utils.api.constants import TEST_DATA_PATH


class DataHelper:
    """
    Класс с полезными методами для работы с данными
    """

    def __init__(self, plastilin_db_api, users_api):
        self.plastilin_db_api = plastilin_db_api
        self.users_api = users_api

    def get_or_create_spec_id_by_name(self, token: str, russian_name: str) -> Optional[int]:
        """
        Ищет spec_id по русскому названию или создает новую запись

        Args:
            token: Bearer токен
            russian_name: Русское название для поиска

        Returns:
            int: spec_id если найден или создан, None если произошла ошибка

        Raises:
            Exception: При ошибке получения или создания культуры
        """
        if not token or not russian_name:
            raise ValueError('Token и russian_name обязательны')
        response = self.plastilin_db_api.get_species_table(token=token)

        if response.status != HTTPStatus.OK:
            raise Exception(f'Ошибка получения списка культур: {response.status}')

        data = response.json()
        species_list = data.get('data', [])

        for item in species_list:
            if item['russian_name'].strip() == russian_name.strip():
                return item['spec_id']

        response = self.plastilin_db_api.create_species_table(
            token=token, russian_name=russian_name
        )

        if response.status == HTTPStatus.CREATED:
            created_data = response.json()
            return created_data['spec_id']
        else:
            raise Exception(f'Ошибка создания культуры: {response.status}')

    def get_my_company_id(self, token: str) -> Optional[int]:
        """
        Получает ID компании для текущего пользователя
        """
        response = self.users_api.get_company_structure(token=token)
        if response.status != HTTPStatus.OK:
            raise Exception(f'Ошибка получения списка компаний: {response.status}')
        return response.json().get('id')

    def generate_random_string(self, name: Optional[str] = None, length: int = 10) -> str:
        """
        Генерирует случайный строковый идентификатор
        """
        if name:
            return f'{name}_{str(uuid.uuid4())[:length]}'
        else:
            return str(uuid.uuid4())[:length]

    def get_user_id(self, token: str) -> Optional[int]:
        """
        Получает ID пользователя для текущего токена
        """
        response = self.users_api.get_profile(token=token)
        return response.json().get('user_data', {}).get('id')

    def get_user_ids_by_usernames(self, get_token, usernames):
        '''
        Возвращает список user_id по списку username (логинов).
        '''
        if not usernames:
            return None
        ids = []
        for username in usernames:
            token = get_token(username)
            ids.append(self.get_user_id(token=token))
        return ids

    def delete_all_field_year_permissions(self, token: str) -> None:
        """
        Удаляет все разрешения на полевые годы для текущего пользователя
        """
        response = self.plastilin_db_api.get_field_year_permissions(token=token)
        if response.status != HTTPStatus.OK:
            raise Exception(f'Ошибка получения разрешений: {response.status}')
        
        data = response.json()
        
        for permission in data:
            permission_id = permission.get('id')
            if permission_id:
                try:
                    response = self.plastilin_db_api.delete_field_year_permissions(
                        token=token, 
                        field_year_permission_id=permission_id
                    )
                except Exception:
                    print(f'Ошибка удаления разрешения: {permission_id}')



    def get_or_create_year_id(
        self, token: str, year: int, spec_id: int, field_id: int, region: str
    ) -> Optional[int]:
        """
        Получает ID года для текущего пользователя
        """
        if not token or not year or not spec_id:
            raise ValueError('Token, year и spec_id обязательны')

        response = self.plastilin_db_api.get_field_year(token=token, spec_id=spec_id)

        if response.status != HTTPStatus.OK:
            raise Exception(f'Ошибка получения списка полевых лет: {response.status}')

        data = response.json()

        year_key = str(year)
        if year_key in data:
            for year_data in data[year_key]:
                if year_data.get('field_id') == field_id:
                    return year_data.get('year_id')

        response = self.plastilin_db_api.add_field_year(
            token=token, year=year, spec_id=spec_id, field_id=field_id
        )

        if response.status == HTTPStatus.CREATED:
            created_data = response.json()
            return created_data.get('id')
        else:
            raise Exception(f'Ошибка создания года: {response.status}')

    def get_or_create_field_id_by_name(
        self, token: str, spec_id: int, field_name: str, region: str
    ) -> Optional[int]:
        """
        Получает ID питомника для текущего пользователя
        """
        if not token or not spec_id or not field_name:
            raise ValueError('Token, spec_id и field_name обязательны')

        response = self.plastilin_db_api.get_field_table(token=token, spec_id=spec_id)

        if response.status != HTTPStatus.OK:
            raise Exception(f'Ошибка получения списка полей: {response.status}')

        data = response.json()
        field_list = data

        for item in field_list:
            if item.get('field_name') == field_name and item.get('region') == region:
                return item.get('field_id')

        response = self.plastilin_db_api.create_field_table(
            token=token, spec_id=spec_id, field_name=field_name, region=region
        )

        if response.status == HTTPStatus.OK:
            created_data = response.json()
            return created_data.get('field_id')
        else:
            raise Exception(f'Ошибка создания питомника: {response.status}')

    def get_or_create_spec_field_year_id(
        self, token: str, spec_name: str, field_name: str, year: int, region: str
    ) -> Tuple[Optional[int], Optional[int], Optional[int]]:
        """
        Получает ID культуры, питомника и года для текущего пользователя
        """
        spec_id = self.get_or_create_spec_id_by_name(token=token, russian_name=spec_name)
        if spec_id is None:
            raise Exception(f'Не удалось получить или создать spec для имени {spec_name}')
        field_id = self.get_or_create_field_id_by_name(
            token=token, spec_id=spec_id, field_name=field_name, region=region
        )
        if field_id is None:
            raise Exception(f'Не удалось получить или создать field для имени {field_name}')
        year_id = self.get_or_create_year_id(
            token=token, year=year, spec_id=spec_id, field_id=field_id, region=region
        )
        if year_id is None:
            raise Exception(f'Не удалось получить или создать year для имени {year}')
        return spec_id, field_id, year_id

    @staticmethod
    def universal_sort_key(s, reverse=False):
        """
        Возвращает ключ для универсальной сортировки строк, чисел и None значений.
        Поддерживает естественную сортировку строк с цифрами.

        Args:
            s: Строка, число или None для сортировки
            reverse: Если True, None значения будут в начале, иначе в конце

        Returns:
            tuple: (буквенная_часть, числовая_часть) для сортировки
        """
        if s is None:
            if reverse:
                return ('', -999999)
            else:
                return ('zzzzzzzzzz', 999999)

        # Обрабатываем числовые значения
        if isinstance(s, (int, float)):
            return ('', s)

        # Нормализуем строку (заменяем ё на е)
        s_norm = s.replace('ё', 'е').replace('Ё', 'Е')

        # # Пока решили не использовать сортировку по цифрам в конце строки
        # m = re.match(r'(\D+?)(\d+)$', s_norm.strip())
        # if m:
        #     prefix, num = m.groups()
        #     num_val = int(num)
        #     if reverse:
        #         return (prefix.lower(), -num_val)
        #     else:
        #         return (prefix.lower(), num_val)

        return (s_norm.lower(), 0)

    def get_multiple_values_from_response(
        self, response_data: list, name_of_multiple: str, feature: str
    ) -> list:
        """
        Извлекает значения из ответа API для фильтрации multiple параметров

        Args:
            response_data: Данные ответа от API
            name_of_multiple: Название поля для извлечения
            feature: Название feature для поиска в features

        Returns:
            list: Список значений для проверки
        """
        if name_of_multiple in ['line_name', 'plot_name']:
            return [item[name_of_multiple] for item in response_data]
        else:
            return [
                next(
                    feat[name_of_multiple]
                    for feat in item['features']
                    if feat.get('feature') == feature
                )
                for item in response_data
            ]

    
    def _generate_field_value(self, field_type: str) -> Any:
        """Генерирует случайное значение для поля по типу"""
        if field_type == 'float':
            value = round(random.uniform(0, 100), 2)
            return int(value) if value.is_integer() else value
        elif field_type == 'date':
            current_year = datetime.now().year
            start_date = datetime(current_year, 1, 1)
            end_date = datetime(current_year, 12, 31)
            days_between = (end_date - start_date).days
            random_days = random.randint(0, days_between)
            random_date = start_date + timedelta(days=random_days)
            return random_date.strftime('%Y-%m-%d')
        return self.generate_random_string('Тестовое значение', length=6)

    def generate_plot_result_data(
        self,
        field_name: str,
        year: int,
        region: str,
        base_plot_name: str,
        number_of_plots: Optional[str] = None,
        sort_name: Optional[bool] = True,
        row_count: int = 10,
        repeats: Optional[int] = None,
        phenotypic_fields: Optional[List[Dict[str, str]]] = None,
        dev_stage_fields: Optional[List[Dict[str, str]]] = None,
    ) -> Dict[str, Any]:
        """
        Генерирует данные для результатов делянок в двух форматах
        
        Args:
            field_name: Название питомника
            year: Год
            region: Регион
            sort_name: Название сорта
            base_plot_name: Базовое название делянки
            row_count: Количество записей на повторность
            repeats: Количество повторностей
            phenotypic_fields: [{'name': 'Высота растения', 'type': 'float', 'unit': 'см'}, ...]
            dev_stage_fields: [{'name': 'Развертывание первых листьев', 'type': 'date'}, ...]
            
        Returns:
            Dict с двумя форматами данных:
            {
                'excel': {
                    'headers': [...],
                    'data': [...],
                    'total_rows': 60,
                    'field_names': [...],
                    'dev_stage_names': [...],
                    'metadata': {...}
                },
                'api': [
                    {
                        'plot_name': 'Делянка 1/1',
                        'line_name': 'Сорт 1',
                        'plot_results': [...],
                        'plot_stages': [...]
                    },
                    ...
                ]
            }
        """
        phenotypic_fields = phenotypic_fields or []
        dev_stage_fields = dev_stage_fields or []
        
        # Формируем заголовки для Excel динамически
        headers = []
        
        # Добавляем базовые заголовки только если соответствующие параметры переданы
        if number_of_plots:
            headers.append('Номер делянки')
        if field_name:
            headers.append('Название питомника')
        if year:
            headers.append('Год')
        if region:
            headers.append('Регион')
        if base_plot_name:
            headers.append('Название делянки')
        if sort_name:
            headers.append('Название сорта')
        if repeats:
            headers.append('Номер повторности')
        
        phenotypic_names = []
        # Формируем заголовки фенотипов
        if phenotypic_fields:
            for field in phenotypic_fields:
                if field['type'] == 'float' and 'unit' in field:
                    phenotypic_names.append(f"Фенотип;{field['name']}; {field['unit']}")
                else:
                    phenotypic_names.append(f"Фенотип;{field['name']};")

        dev_stage_names = []
        if dev_stage_fields:
            for field in dev_stage_fields:
                dev_stage_names.append(f"Стадия развития;{field['name']};")
        
        headers.extend(phenotypic_names)
        headers.extend(dev_stage_names)
        
        # Генерируем данные для обоих форматов
        excel_data = []
        api_data = []

        count_of_plots = 0

        for repeat in range(1, repeats + 1 if repeats else 2):
            for plot_num in range(1, row_count + 1):
                # Генерируем значения для текущей записи
                
                if phenotypic_names:
                    phenotypic_values = [
                        self._generate_field_value(f['type']) for f in phenotypic_fields
                    ]
                if dev_stage_names:
                    dev_stage_values = [
                        self._generate_field_value(f['type']) for f in dev_stage_fields
                    ]
                
                # Формируем строку для Excel динамически
                excel_row = []
                
                # Добавляем данные только если соответствующие параметры переданы
                if number_of_plots:
                    excel_row.append(f'{number_of_plots} {count_of_plots + 1}')
                    count_of_plots += 1
                if field_name:
                    excel_row.append(field_name)
                if year:
                    excel_row.append(year)
                if region:
                    excel_row.append(region)
                if base_plot_name:
                    excel_row.append(f"{base_plot_name} {plot_num}")
                if sort_name:
                    excel_row.append(f"Сорт {plot_num}")
                if repeats:
                    excel_row.append(repeat)
                if phenotypic_names:
                    excel_row.extend(phenotypic_values)
                if dev_stage_names:
                    excel_row.extend(dev_stage_values)
                excel_data.append(excel_row)
                
                # Формируем данные для API
                plot_results = []
                if phenotypic_fields:
                    for i, field in enumerate(phenotypic_fields):
                        plot_results.append({
                            "plot_final_feature": field['name'].lower(),
                            "plot_final_value": str(phenotypic_values[i]),
                            "plot_final_unit": field.get('unit', '')
                        })
                
                plot_stages = []
                if dev_stage_fields:
                    for i, field in enumerate(dev_stage_fields):
                        plot_stages.append({
                            "stage_of_vegetation": field['name'].lower(),
                            "date_of_stage": dev_stage_values[i]
                    })
                
                api_record = {
                    "plot_name": f"{base_plot_name} {plot_num}/{repeat}",
                    "line_name": f"Сорт {plot_num}",
                    "plot_label": f"{number_of_plots} {repeat}" if number_of_plots else None,
                    "plot_results": plot_results if phenotypic_fields else None,
                    "plot_stages": plot_stages if dev_stage_fields else None
                }
                api_data.append(api_record)
        
        return {
            'excel': {
                'headers': headers,
                'data': excel_data,
                'total_rows': len(excel_data),
                'field_names': phenotypic_names,
                'dev_stage_names': dev_stage_names,
                'metadata': {
                    'field_name': field_name,
                    'year': year,
                    'region': region,
                    'base_plot_name': base_plot_name,
                    'row_count': row_count,
                    'repeats': repeats
                }
            },
            'api': api_data
        }

    def update_plot_result_data(
        self,
        plot_data: Dict[str, Any],
        modifications: Optional[List[Dict[str, Any]]] = None,
        add_columns: Optional[List[Dict[str, Any]]] = None,
        remove_columns: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """
        Обновляет данные в обоих форматах (Excel и API)
        
        Args:
            plot_data: Данные из generate_plot_result_data
            modifications: Список изменений значений
                [
                    {
                        'plot_name': 'Делянка 1/1',
                        'line_name': 'Сорт 1',
                        'plot_results': [
                            {
                                'plot_final_feature': 'высота растения',
                                'plot_final_value': '150.5',
                                'plot_final_unit': 'см'
                            }
                        ],
                        'plot_stages': [
                            {
                                'stage_of_vegetation': 'всходы',
                                'date_of_stage': '2024-05-15'
                            }
                        ]
                    }
                ]
            add_columns: Список новых колонок для добавления
                [
                    {
                        'excel_header': 'Фенотип;Новый признак; шт',
                        'api_field': 'новый признак',
                        'api_unit': 'шт',
                        'type': 'float',
                        'default_value': 10.0
                    }
                ]
            remove_columns: Список колонок для удаления
                ['Фенотип;Устойчивость к болезням;']
        """
        updated_data = plot_data.copy()
        
        # Обновляем Excel данные
        excel_data = updated_data['excel'].copy()
        excel_data['data'] = [row.copy() for row in excel_data['data']]
        
        # Обновляем API данные
        api_data = updated_data['api'].copy()
        
        # Удаляем колонки
        if remove_columns:
            for col_to_remove in remove_columns:
                # Удаляем из Excel
                if col_to_remove in excel_data['headers']:
                    col_index = excel_data['headers'].index(col_to_remove)
                    excel_data['headers'].pop(col_index)
                    for row in excel_data['data']:
                        row.pop(col_index)
                
                # Удаляем из API
                for record in api_data:
                    # Удаляем из plot_results
                    record['plot_results'] = [
                        result for result in record['plot_results']
                        if (result['plot_final_feature'] != 
                            col_to_remove.split(';')[1].strip().lower())
                    ]
                    # Удаляем из plot_stages
                    record['plot_stages'] = [
                        stage for stage in record['plot_stages']
                        if (stage['stage_of_vegetation'] != 
                            col_to_remove.split(';')[1].strip().lower())
                    ]
        
        # Добавляем колонки
        if add_columns:
            for new_col in add_columns:
                excel_header = new_col['excel_header']
                api_field = new_col['api_field']
                api_unit = new_col.get('api_unit', '')
                col_type = new_col.get('type', 'string')
                default_value = new_col.get('default_value', '')
                
                # Добавляем в Excel
                excel_data['headers'].append(excel_header)
                for row in excel_data['data']:
                    if col_type == 'float':
                        row.append(default_value)
                    else:
                        row.append(default_value)
                
                # Добавляем в API
                for record in api_data:
                    if excel_header.startswith('Фенотип;'):
                        record['plot_results'].append({
                            "plot_final_feature": api_field,
                            "plot_final_value": str(default_value),
                            "plot_final_unit": api_unit
                        })
                    elif excel_header.startswith('Стадия развития;'):
                        record['plot_stages'].append({
                            "stage_of_vegetation": api_field,
                            "date_of_stage": default_value
                        })
        
        # Обновляем значения
        if modifications:
            for mod in modifications:
                plot_name = mod.get('plot_name', '')
                line_name = mod.get('line_name', '')
                new_plot_results = mod.get('plot_results', [])
                new_plot_stages = mod.get('plot_stages', [])
                
                # Находим соответствующую запись в API данных
                api_record = None
                for record in api_data:
                    if (record['plot_name'] == plot_name and 
                        record['line_name'] == line_name):
                        api_record = record
                        break
                
                if api_record:
                    # Обновляем plot_results
                    if new_plot_results:
                        for new_result in new_plot_results:
                            feature = new_result.get('plot_final_feature', '')
                            # Ищем существующий результат
                            existing_result = None
                            for result in api_record['plot_results']:
                                if result['plot_final_feature'] == feature:
                                    existing_result = result
                                    break
                            
                            if existing_result:
                                # Обновляем существующий
                                existing_result.update(new_result)
                            else:
                                # Добавляем новый
                                api_record['plot_results'].append(new_result)
                    
                    # Обновляем plot_stages
                    if new_plot_stages:
                        for new_stage in new_plot_stages:
                            stage_name = new_stage.get('stage_of_vegetation', '')
                            # Ищем существующую стадию
                            existing_stage = None
                            for stage in api_record['plot_stages']:
                                if stage['stage_of_vegetation'] == stage_name:
                                    existing_stage = stage
                                    break
                            
                            if existing_stage:
                                # Обновляем существующую
                                existing_stage.update(new_stage)
                            else:
                                # Добавляем новую
                                api_record['plot_stages'].append(new_stage)
                
                # Парсим plot_name для извлечения названия делянки и повторности
                # Формат: "Делянка 1/1" -> название: "Делянка 1", повторность: 1
                plot_base_name = plot_name
                repeat_number = 1
                if '/' in plot_name:
                    parts = plot_name.split('/')
                    if len(parts) == 2:
                        plot_base_name = parts[0].strip()
                        try:
                            repeat_number = int(parts[1].strip())
                        except ValueError:
                            repeat_number = 1
                
                # Находим соответствующую строку в Excel данных
                excel_row_index = None
                for i, row in enumerate(excel_data['data']):
                    # plot_name находится в 4-й колонке (индекс 3)
                    # line_name находится в 5-й колонке (индекс 4) 
                    # номер повторности находится в 6-й колонке (индекс 5)
                    if (row[3] == plot_base_name and 
                        row[4] == line_name and
                        row[5] == repeat_number):
                        excel_row_index = i
                        break
                
                if excel_row_index is not None:
                    # Обновляем Excel данные на основе обновленных API данных
                    updated_record = None
                    for record in api_data:
                        if (record['plot_name'] == plot_name and 
                            record['line_name'] == line_name):
                            updated_record = record
                            break
                    
                    if updated_record:
                        # Обновляем фенотипы в Excel
                        for result in updated_record['plot_results']:
                            feature = result['plot_final_feature']
                            value = result['plot_final_value']
                            
                            # Ищем соответствующую колонку в Excel
                            for i, header in enumerate(excel_data['headers']):
                                if header.startswith('Фенотип;') and feature in header:
                                    excel_data['data'][excel_row_index][i] = value
                                    break
                        
                        # Обновляем стадии развития в Excel
                        for stage in updated_record['plot_stages']:
                            stage_name = stage['stage_of_vegetation']
                            date_value = stage['date_of_stage']
                            
                            # Ищем соответствующую колонку в Excel
                            for i, header in enumerate(excel_data['headers']):
                                if header.startswith('Стадия развития;') and stage_name in header:
                                    excel_data['data'][excel_row_index][i] = date_value
                                    break
        
        updated_data['excel'] = excel_data
        updated_data['api'] = api_data
        
        return updated_data

    def create_plot_result_excel(
        self,
        plot_data: Dict[str, Any],
    ) -> Tuple[Path, str]:
        """
        Создает Excel файл для результатов делянок
        
        Args:
            plot_data: Готовые данные (если None, генерируются автоматически)
            field_name, year, region, base_plot_name: Параметры для генерации данных
            row_count, repeats, phenotypic_fields, dev_stage_fields: Параметры для генерации
        """
        
        # Извлекаем Excel данные
        excel_data = plot_data['excel']
        
        # Создаем Excel
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet('Результаты')
        else:
            ws.title = 'Результаты'
        
        # Записываем заголовки
        for col, header in enumerate(excel_data['headers'], 1):
            ws.cell(row=1, column=col, value=header)
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = 20
        
        # Записываем данные
        for row_data in excel_data['data']:
            ws.append(row_data)
        
        # Сохраняем файл
        filename = (
            f"{excel_data['metadata']['base_plot_name']}_"
            f"{excel_data['metadata']['field_name']}_"
            f"{excel_data['metadata']['year']}_"
            f"{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        )
        # TODO: remove this
        # path = Path(tempfile.gettempdir()) / filename
        path = TEST_DATA_PATH / filename
        wb.save(path)
        wb.close()
        
        return path, filename

    # Сравнивает данные plot_response и plot_data['api']
    def compare_plot_response_with_api(self, response_data, plot_data):
        # Проходим по каждому ответу
        for plot_response in response_data['data']:
            # Ищем соответствующий plot_api по имени
            plot_api = next(
                (p for p in plot_data['api'] if p['plot_name'] == plot_response['plot_name']),
                None,
            )
            assert plot_api is not None
            assert plot_response['line_name'] == plot_api['line_name']

            # Словари для быстрого поиска по признаку



            # Проверяем plot_results только если они есть в plot_data
            if plot_api['plot_results']:
                plot_results_by_feature = {
                    r['plot_final_feature']: r for r in plot_api['plot_results']
                }
                response_results_by_feature = {
                    r['plot_final_feature']: r for r in plot_response['plot_results']
                }
                for feature, plot_result in plot_results_by_feature.items():
                    plot_response_result = response_results_by_feature.get(feature)
                    assert plot_response_result is not None
                    # Сравниваем значения как числа
                    assert (
                        plot_result['plot_final_value'] == 
                        plot_response_result['plot_final_value']
                    ), (f'{plot_result["plot_final_value"]} != '
                        f'{plot_response_result["plot_final_value"]}')
                    assert (
                        plot_result['plot_final_unit'] == 
                        plot_response_result['plot_final_unit']
                    ), (f'{plot_result["plot_final_unit"]} != '
                        f'{plot_response_result["plot_final_unit"]}')

            # Аналогично для стадий
            

            # Проверяем plot_stages только если они есть в plot_data
            if plot_api['plot_stages']:
                plot_stages_by_name = {
                    s['stage_of_vegetation'].lower(): s for s in plot_api['plot_stages']
                }
                response_stages_by_name = {
                    s['stage_of_vegetation'].lower(): s for s in plot_response['plot_stages']
                }
                # Находим дату стадии "всходы"
                sowing_stage = plot_stages_by_name.get('всходы')
                assert sowing_stage is not None
                sowing_date = sowing_stage['date_of_stage']

                for stage_name, plot_stage in plot_stages_by_name.items():
                    response_stage = response_stages_by_name.get(stage_name)
                    assert response_stage is not None
                    assert plot_stage['date_of_stage'] == response_stage['date_of_stage']
                    days_after_sowing = (
                        datetime.strptime(response_stage['date_of_stage'], '%Y-%m-%d') -
                        datetime.strptime(sowing_date, '%Y-%m-%d')
                    ).days
                    assert int(days_after_sowing) == int(response_stage['days_after_sowing'])

    def generate_random_plot_data(self, plot_fields: Dict[str, Any], spec_name: str, token: str):
        
        plot_data = self.generate_plot_result_data(
            field_name=plot_fields['field_name'],
            year=plot_fields['year'],
            region=plot_fields['region'],
            base_plot_name=plot_fields['base_plot_name'],
            number_of_plots=plot_fields['number_of_plots'],
            sort_name=plot_fields['sort_name'],
            row_count=plot_fields['row_count'],
            repeats=plot_fields['repeats'],
            phenotypic_fields=plot_fields['phenotypic_fields'],
            dev_stage_fields=plot_fields['dev_stage_fields'],
        )
        
        spec_id, field_id, year_id = self.get_or_create_spec_field_year_id(
            token=token,
            spec_name=spec_name,
            field_name=plot_fields['field_name'],
            year=plot_fields['year'],
            region=plot_fields['region'],
        )

        excel_file, filename = self.create_plot_result_excel(plot_data=plot_data)

        response = self.plastilin_db_api.create_plot_result(
            token=token,
            file_name=filename,
            file_path=excel_file,
            spec_id=spec_id,
        )
        assert response.status == HTTPStatus.CREATED

        return {
            'spec_id': spec_id,
            'field_id': field_id,
            'year_id': year_id,
        }